import subprocess, sys

def install(pkg):
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

try:
    import openpyxl
except ImportError:
    install("openpyxl")
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colour palette ──────────────────────────────────────────────────────────
C_HEADER_BG   = "3B4CB8"   # indigo
C_HEADER_FG   = "FFFFFF"
C_SUBHEAD_BG  = "E8EAFD"
C_SUBHEAD_FG  = "1A237E"
C_ALT_ROW     = "F5F6FF"
C_MUST        = "D32F2F"
C_SHOULD      = "F57C00"
C_COULD       = "388E3C"
C_HIGH_RISK   = "FFCDD2"
C_MED_RISK    = "FFE0B2"
C_LOW_RISK    = "F1F8E9"
C_FREE_BADGE  = "E3F2FD"
C_PRO_BADGE   = "FFF3E0"
C_BORDER      = "C5CAE9"

def header_style(ws, row, cols, text_list, bg=C_HEADER_BG, fg=C_HEADER_FG, size=11):
    for i, txt in enumerate(text_list, 1):
        cell = ws.cell(row=row, column=i, value=txt)
        cell.font = Font(bold=True, color=fg, size=size)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color=C_BORDER)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def data_row(ws, row, values, alt=False, height=None):
    fill = PatternFill("solid", fgColor=C_ALT_ROW) if alt else None
    thin = Side(style="thin", color=C_BORDER)
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if height:
        ws.row_dimensions[row].height = height

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def priority_color(p):
    if p == "Must":    return C_MUST
    if p == "Should":  return C_SHOULD
    if p == "Could":   return C_COULD
    return "000000"

def color_cell(cell, hex_color):
    thin = Side(style="thin", color=C_BORDER)
    cell.font = Font(bold=True, color=hex_color)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.alignment = Alignment(horizontal="center", vertical="center")


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 — Overview
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Overview"
ws.sheet_view.showGridLines = False

header_style(ws, 1, 2, ["AI Chat Mobile App — PRD Summary", ""], size=14)
ws.merge_cells("A1:B1")
ws.row_dimensions[1].height = 36

rows = [
    ("Document Version", "1.0"),
    ("Date", "June 2026"),
    ("Status", "Draft"),
    ("Platforms", "iOS 16+ · Android API 29+"),
    ("Tech Stack (Mobile)", "Flutter (Dart) · Riverpod / BLoC"),
    ("Tech Stack (Backend)", "Python 3.12 · FastAPI · PostgreSQL · Redis"),
    ("AI Providers", "Anthropic Claude · OpenAI GPT · Google Gemini"),
    ("Monetization", "Freemium + Pro $9.99/month"),
    ("Free Tier Limit", "20 messages/day · Haiku, GPT-4o-mini, Gemini Flash"),
    ("Pro Tier", "Unlimited messages · All models · Image/File uploads · Voice"),
    ("Target Users", "General consumers — Maya (writer), Arjun (engineer), Zoe (student)"),
    ("MVP Timeline", "Months 1–3"),
    ("Launch Timeline", "Month 6"),
]

for r, (k, v) in enumerate(rows, 2):
    ws.cell(row=r, column=1, value=k).font = Font(bold=True, color=C_SUBHEAD_FG)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=C_SUBHEAD_BG)
    ws.cell(row=r, column=1).alignment = Alignment(vertical="center")
    ws.cell(row=r, column=2, value=v).alignment = Alignment(vertical="center", wrap_text=True)
    if r % 2 == 1:
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=C_ALT_ROW)
    ws.row_dimensions[r].height = 20

set_col_widths(ws, [28, 60])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 — Goals & Metrics
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Goals & Metrics")
ws2.sheet_view.showGridLines = False

header_style(ws2, 1, 4, ["Goal", "KPI", "Target (6 months)", "Category"])
ws2.row_dimensions[1].height = 28

goals = [
    ("User Acquisition",    "Total registered users",       "50,000",          "Business"),
    ("Retention",           "D30 retention rate",            "≥ 25%",           "Business"),
    ("Monetization",        "Pro conversion rate",           "≥ 5% of MAU",     "Business"),
    ("Revenue",             "Monthly Recurring Revenue",     "$25,000",         "Business"),
    ("App Quality",         "App Store / Play Store rating", "≥ 4.4 stars",     "Business"),
    ("Onboarding Speed",    "Time to first message",         "≤ 90 seconds",    "Product"),
    ("Streaming Latency",   "Token-to-render latency",       "≤ 100 ms",        "Product"),
    ("Security",            "AI API key leaks to client",    "0 incidents",     "Product"),
    ("Availability",        "Backend uptime",                "99.5%",           "Product"),
    ("First Token",         "First token latency (p95)",     "< 1.5 seconds",   "Product"),
    ("Cold Start",          "App cold start (mid-range)",    "< 3 seconds",     "Product"),
]

for i, row in enumerate(goals):
    alt = i % 2 == 1
    data_row(ws2, i + 2, row, alt=alt, height=20)

set_col_widths(ws2, [28, 36, 24, 16])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3 — User Stories
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("User Stories")
ws3.sheet_view.showGridLines = False

header_style(ws3, 1, 6, ["ID", "Area", "As a…", "I want to…", "So that…", "Priority"])
ws3.row_dimensions[1].height = 28

stories = [
    # Auth
    ("US-01","Auth","New user","Sign up with email and password","I can create an account","Must"),
    ("US-02","Auth","New user","Sign up with Google or Apple","I can skip form filling","Must"),
    ("US-03","Auth","Returning user","Log in and stay logged in","I don't re-enter credentials each session","Must"),
    ("US-04","Auth","User","Reset my password via email","I can recover access","Must"),
    ("US-05","Auth","User","Delete my account and all data","I can exercise my right to erasure","Must"),
    # Conversations
    ("US-06","Conversations","User","Start a new conversation","I can talk about a new topic","Must"),
    ("US-07","Conversations","User","See all my past conversations in a list","I can continue where I left off","Must"),
    ("US-08","Conversations","User","Rename a conversation","I can identify it later","Should"),
    ("US-09","Conversations","User","Delete a conversation","I can clean up clutter","Must"),
    ("US-10","Conversations","User","Search conversations by keyword","I can find a specific thread","Should"),
    ("US-11","Conversations","User","Pin/star a conversation","Important chats are always at the top","Should"),
    ("US-12","Conversations","User","Archive a conversation","I can hide old threads without deleting","Could"),
    ("US-13","Conversations","User","Access conversations on a new device","History is always in sync","Must"),
    # Messaging
    ("US-14","Messaging","User","Send a message and see a streaming response","The AI feels fast and responsive","Must"),
    ("US-15","Messaging","User","See formatted responses (markdown, code blocks)","Technical content is readable","Must"),
    ("US-16","Messaging","User","Copy a code block with one tap","I can use code quickly","Must"),
    ("US-17","Messaging","User","Regenerate the last AI response","I can get a different answer","Must"),
    ("US-18","Messaging","User","Edit my sent message and re-submit","I can fix a typo without restarting","Should"),
    ("US-19","Messaging","User","Stop a response mid-stream","I can interrupt a long wrong answer","Must"),
    ("US-20","Messaging","User","Copy any message to clipboard","I can share or use the content elsewhere","Should"),
    ("US-21","Messaging","User","Share a conversation as text or screenshot","I can share useful AI answers","Could"),
    # Models
    ("US-22","Models","User","Switch the AI model for a conversation","I can choose the best model for my task","Must"),
    ("US-23","Models","User","See which model is active at a glance","I'm not confused about who I'm talking to","Must"),
    ("US-24","Models","User","Set a default model in settings","New conversations start with my preference","Should"),
    ("US-25","Models","User","See model capability badges","I know if a model supports images before uploading","Should"),
    # Multimodal
    ("US-26","Multimodal","Pro user","Upload an image from gallery or camera","I can ask questions about visuals","Must"),
    ("US-27","Multimodal","Pro user","Upload a PDF or document","I can get summaries or analysis","Must"),
    ("US-28","Multimodal","User","Record a voice message as my input","I can chat hands-free","Should"),
    ("US-29","Multimodal","User","Hear AI responses read aloud","I can consume content without reading","Could"),
    ("US-30","Multimodal","User","Preview attachments before sending","I can confirm I'm sending the right file","Must"),
    # Custom Instructions
    ("US-31","Instructions","User","Set global instructions applied to every chat","The AI always knows my context/preferences","Should"),
    ("US-32","Instructions","User","Override the system prompt per conversation","I can give different personas per context","Should"),
    ("US-33","Instructions","User","Save named prompt templates","I can quickly apply a Code Review or Tutor persona","Could"),
    # Settings
    ("US-34","Settings","User","Switch between dark and light mode","I can use the app comfortably","Must"),
    ("US-35","Settings","User","Change the font size","The text is comfortable to read","Should"),
    ("US-36","Settings","User","Set my preferred response language","The AI replies in my language","Could"),
    # Monetization
    ("US-37","Monetization","Free user","See how many messages I have left today","I know when I'll hit the limit","Must"),
    ("US-38","Monetization","User","Upgrade to Pro from within the app","I can unlock all features seamlessly","Must"),
    ("US-39","Monetization","Pro user","Manage or cancel my subscription","I'm in control of billing","Must"),
    ("US-40","Monetization","User","Restore a previous purchase","My Pro status survives reinstalls","Must"),
    # Privacy
    ("US-41","Privacy","User","Export all my data as JSON or plain text","I have a copy of my conversations","Must"),
    ("US-42","Privacy","User","Delete all my conversation history","I can wipe my data without deleting the account","Must"),
    ("US-43","Privacy","User","Toggle server-side history saving off","My conversations are never stored on the server","Should"),
]

thin = Side(style="thin", color=C_BORDER)
for i, row in enumerate(stories):
    r = i + 2
    alt = i % 2 == 1
    data_row(ws3, r, row, alt=alt, height=22)
    # Color-code priority cell
    pcell = ws3.cell(row=r, column=6)
    color_cell(pcell, priority_color(row[5]))
    ws3.row_dimensions[r].height = 22

set_col_widths(ws3, [8, 16, 14, 40, 42, 10])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4 — Functional Requirements
# ═══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Functional Requirements")
ws4.sheet_view.showGridLines = False

header_style(ws4, 1, 4, ["ID", "Area", "Requirement", "Tier"])
ws4.row_dimensions[1].height = 28

func_reqs = [
    ("FR-01","Auth","Email/password registration and login","Free + Pro"),
    ("FR-02","Auth","Social login — Google and Apple Sign-In","Free + Pro"),
    ("FR-03","Auth","JWT session management with refresh tokens","Free + Pro"),
    ("FR-04","Auth","Password reset via email","Free + Pro"),
    ("FR-05","Auth","Account deletion with full data erasure (GDPR)","Free + Pro"),
    ("FR-06","Auth","User profile (display name, avatar)","Free + Pro"),
    ("FR-07","Conversations","Create, rename, and delete conversation threads","Free + Pro"),
    ("FR-08","Conversations","Conversation list sorted by recency with pinned first","Free + Pro"),
    ("FR-09","Conversations","Full conversation history synced across devices","Free + Pro"),
    ("FR-10","Conversations","Search conversations by keyword (FTS)","Free + Pro"),
    ("FR-11","Conversations","Pin/star important conversations","Free + Pro"),
    ("FR-12","Conversations","Archive conversations","Free + Pro"),
    ("FR-13","Messaging","Send and receive text messages with streaming (SSE)","Free + Pro"),
    ("FR-14","Messaging","Markdown rendering (code, tables, lists, bold, italic)","Free + Pro"),
    ("FR-15","Messaging","Syntax highlighting in code blocks with copy button","Free + Pro"),
    ("FR-16","Messaging","Regenerate last AI response","Free + Pro"),
    ("FR-17","Messaging","Edit a sent message and re-submit","Free + Pro"),
    ("FR-18","Messaging","Stop/cancel an in-progress AI response","Free + Pro"),
    ("FR-19","Messaging","Copy individual messages to clipboard","Free + Pro"),
    ("FR-20","Messaging","Share conversation as text or image","Free + Pro"),
    ("FR-21","Models","Per-conversation model selector","Free + Pro"),
    ("FR-22","Models","Display active model name in conversation header","Free + Pro"),
    ("FR-23","Models","Model capabilities badge (images, files, speed)","Free + Pro"),
    ("FR-24","Models","Default model preference in user settings","Free + Pro"),
    ("FR-25","Multimodal","Image upload — camera capture or gallery picker","Pro"),
    ("FR-26","Multimodal","File upload — PDF, DOCX, TXT (up to 25 MB)","Pro"),
    ("FR-27","Multimodal","Voice input — speech-to-text for message composition","Pro"),
    ("FR-28","Multimodal","Voice output — TTS for AI responses (toggleable)","Pro"),
    ("FR-29","Multimodal","Preview uploaded files/images before sending","Pro"),
    ("FR-30","Multimodal","Multiple attachments per message","Pro"),
    ("FR-31","Instructions","Global custom instructions for every conversation","Free + Pro"),
    ("FR-32","Instructions","Per-conversation system prompt override","Free + Pro"),
    ("FR-33","Instructions","Saved prompt templates (Personas)","Free + Pro"),
    ("FR-34","Settings","Dark / light / system default theme","Free + Pro"),
    ("FR-35","Settings","Font size adjustment","Free + Pro"),
    ("FR-36","Settings","Response language preference","Free + Pro"),
    ("FR-37","Settings","Notification preferences","Free + Pro"),
    ("FR-38","Settings","App language (i18n)","Free + Pro"),
    ("FR-39","Privacy","Export all conversation data as JSON or plain text","Free + Pro"),
    ("FR-40","Privacy","Delete individual conversations or all history","Free + Pro"),
    ("FR-41","Privacy","Toggle whether conversation history is saved to server","Free + Pro"),
    ("FR-42","Privacy","View data usage and token consumption history","Free + Pro"),
    ("FR-43","Monetization","Free tier — limited messages per day per model","Free"),
    ("FR-44","Monetization","Pro tier — unlimited messages, advanced models","Pro"),
    ("FR-45","Monetization","In-app purchase via App Store and Google Play","Pro"),
    ("FR-46","Monetization","Subscription management screen","Pro"),
    ("FR-47","Monetization","Free tier usage meter visible in UI","Free"),
]

for i, row in enumerate(func_reqs):
    alt = i % 2 == 1
    data_row(ws4, i + 2, row, alt=alt, height=20)
    tcell = ws4.cell(row=i + 2, column=4)
    if row[3] == "Pro":
        tcell.fill = PatternFill("solid", fgColor=C_PRO_BADGE)
        tcell.font = Font(bold=True, color="E65100")
    else:
        tcell.fill = PatternFill("solid", fgColor=C_FREE_BADGE)
        tcell.font = Font(bold=True, color="0D47A1")
    tcell.alignment = Alignment(horizontal="center", vertical="center")

set_col_widths(ws4, [10, 18, 56, 14])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 5 — Non-Functional Requirements
# ═══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Non-Functional Reqs")
ws5.sheet_view.showGridLines = False

header_style(ws5, 1, 4, ["ID", "Category", "Requirement", "Target / Detail"])
ws5.row_dimensions[1].height = 28

nfr = [
    ("NFR-01","Performance","First meaningful paint","< 2 seconds on mid-range device"),
    ("NFR-02","Performance","Streaming token render latency","< 100 ms after receipt"),
    ("NFR-03","Performance","App cold start","< 3 seconds"),
    ("NFR-04","Performance","Offline mode","Read cached conversations without network"),
    ("NFR-05","Scalability","Horizontal scaling","Stateless FastAPI workers"),
    ("NFR-06","Scalability","Streaming transport","Async SSE (Server-Sent Events)"),
    ("NFR-07","Scalability","Database design","Millions of conversation rows supported"),
    ("NFR-08","Security","AI API key storage","Backend only — never in mobile app"),
    ("NFR-09","Security","Transport security","HTTPS / TLS 1.3"),
    ("NFR-10","Security","Password hashing","Argon2id (64 MB memory, 3 iterations)"),
    ("NFR-11","Security","Rate limiting","All API endpoints; 60 req/min on auth"),
    ("NFR-12","Security","Input sanitization","Prompt injection mitigation at API layer"),
    ("NFR-13","Security","JWT tokens","15-min access token; refresh rotated on use"),
    ("NFR-14","Compliance","GDPR","Right to erasure + data export"),
    ("NFR-15","Compliance","CCPA","Opt-out of data sale"),
    ("NFR-16","Compliance","App Store & Play Store","Privacy labels accurately filled"),
    ("NFR-17","Accessibility","Contrast ratio","WCAG AA in both light and dark themes"),
    ("NFR-18","Accessibility","Screen reader","TalkBack (Android) / VoiceOver (iOS)"),
    ("NFR-19","Accessibility","Touch target size","Minimum 44×44 pt"),
    ("NFR-20","Availability","API backend","99.5% monthly uptime SLA"),
    ("NFR-21","Availability","Database","99.9% — managed PostgreSQL + failover"),
]

for i, row in enumerate(nfr):
    data_row(ws5, i + 2, row, alt=i % 2 == 1, height=20)

set_col_widths(ws5, [10, 18, 40, 44])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 6 — Model & Tier Matrix
# ═══════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Models & Tiers")
ws6.sheet_view.showGridLines = False

header_style(ws6, 1, 6, ["Provider", "Model", "Speed", "Capability", "Free Tier", "Pro Tier"])
ws6.row_dimensions[1].height = 28

models = [
    ("Anthropic", "Claude Haiku 4.5",    "Fast",     "Text",                 "✓", "✓"),
    ("Anthropic", "Claude Sonnet 4.6",   "Balanced", "Text + Images + Files", "✗", "✓"),
    ("Anthropic", "Claude Opus 4.8",     "Powerful", "Text + Images + Files", "✗", "✓"),
    ("OpenAI",    "GPT-4o mini",         "Fast",     "Text",                 "✓", "✓"),
    ("OpenAI",    "GPT-4o",              "Balanced", "Text + Images + Files", "✗", "✓"),
    ("Google",    "Gemini 1.5 Flash",    "Fastest",  "Text + Images",        "✓", "✓"),
    ("Google",    "Gemini 1.5 Pro",      "Powerful", "Text + Images + Files", "✗", "✓"),
]

for i, row in enumerate(models):
    r = i + 2
    alt = i % 2 == 1
    data_row(ws6, r, row, alt=alt, height=22)
    for col, val in [(5, row[4]), (6, row[5])]:
        cell = ws6.cell(row=r, column=col)
        if val == "✓":
            cell.font = Font(bold=True, color="2E7D32")
        else:
            cell.font = Font(bold=True, color="C62828")
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Tier comparison table
ws6.cell(row=10, column=1, value="Feature Comparison by Tier").font = Font(bold=True, size=12, color=C_SUBHEAD_FG)
ws6.row_dimensions[10].height = 24

header_style(ws6, 11, 3, ["Feature", "Free", "Pro ($9.99/month)"])
ws6.row_dimensions[11].height = 28

tier_rows = [
    ("Messages per day",        "20",         "Unlimited"),
    ("Models available",        "Haiku, Flash, GPT-4o-mini", "All models incl. Opus, GPT-4o, Gemini Pro"),
    ("Image uploads",           "✗",          "✓ (up to 5/message)"),
    ("File uploads",            "✗",          "✓ (up to 3/message)"),
    ("Voice I/O",               "✗",          "✓"),
    ("Conversation history",    "30 days",    "Unlimited"),
    ("Custom instructions",     "✓",          "✓"),
    ("Priority queue",          "✗",          "✓"),
]

for i, row in enumerate(tier_rows):
    r = i + 12
    data_row(ws6, r, row, alt=i % 2 == 1, height=20)
    for col in [2, 3]:
        cell = ws6.cell(row=r, column=col)
        v = cell.value
        if v == "✓":
            cell.font = Font(bold=True, color="2E7D32")
        elif v == "✗":
            cell.font = Font(bold=True, color="C62828")
        cell.alignment = Alignment(horizontal="center", vertical="center")

set_col_widths(ws6, [16, 28, 16, 32, 10, 10])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 7 — API Endpoints
# ═══════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("API Endpoints")
ws7.sheet_view.showGridLines = False

header_style(ws7, 1, 5, ["Method", "Path", "Description", "Auth Required", "Notes"])
ws7.row_dimensions[1].height = 28

apis = [
    # Auth
    ("POST",   "/auth/register",                           "Register with email + password",                      "No",  "Returns access + refresh tokens"),
    ("POST",   "/auth/login",                              "Login; returns access + refresh tokens",               "No",  ""),
    ("POST",   "/auth/logout",                             "Revoke refresh token",                                "Yes", ""),
    ("POST",   "/auth/refresh",                            "Exchange refresh token for new access token",          "No",  "Rotates refresh token"),
    ("POST",   "/auth/forgot-password",                    "Send password reset email",                           "No",  "Always returns 200 (no email enumeration)"),
    ("POST",   "/auth/reset-password",                     "Apply new password with token",                       "No",  "Token valid 1 hour, single use"),
    ("POST",   "/auth/oauth/google",                       "Google OAuth sign-in",                                "No",  ""),
    ("POST",   "/auth/oauth/apple",                        "Apple Sign-In",                                       "No",  "Required on iOS"),
    # Conversations
    ("GET",    "/conversations",                           "List conversations (paginated)",                       "Yes", "page, limit query params"),
    ("POST",   "/conversations",                           "Create new conversation",                              "Yes", ""),
    ("GET",    "/conversations/{id}",                      "Get conversation + messages",                          "Yes", ""),
    ("PATCH",  "/conversations/{id}",                      "Update title, pinned, archived",                       "Yes", ""),
    ("DELETE", "/conversations/{id}",                      "Delete conversation",                                  "Yes", ""),
    ("GET",    "/conversations/search",                    "Full-text search (q= param)",                          "Yes", "FTS via PostgreSQL tsvector"),
    # Messages
    ("POST",   "/conversations/{id}/messages",             "Send message; returns SSE stream",                     "Yes", "text/event-stream response"),
    ("DELETE", "/conversations/{id}/messages/current-stream", "Cancel active stream",                              "Yes", ""),
    ("PATCH",  "/conversations/{id}/messages/{msg_id}",   "Edit message",                                         "Yes", "Truncates subsequent messages"),
    ("DELETE", "/conversations/{id}/messages/{msg_id}",   "Delete message",                                       "Yes", ""),
    # Uploads
    ("POST",   "/upload/image",                            "Upload image; returns attachment_id",                  "Yes", "Max 20 MB; JPEG, PNG, WEBP, GIF"),
    ("POST",   "/upload/file",                             "Upload document; returns attachment_id",               "Yes", "Max 25 MB; PDF, DOCX, TXT, CSV, XLSX"),
    # User
    ("GET",    "/users/me",                                "Get current user profile + settings",                  "Yes", ""),
    ("PATCH",  "/users/me",                                "Update display name, avatar, settings",                "Yes", ""),
    ("DELETE", "/users/me",                                "Delete account (GDPR erasure)",                        "Yes", "Completes within 30 days"),
    ("GET",    "/users/me/usage",                          "Daily message count + limits",                         "Yes", ""),
    ("POST",   "/users/me/export",                         "Request data export (async)",                          "Yes", "Email with download link within 10 min"),
    # Subscription
    ("POST",   "/subscription/verify",                     "Verify RevenueCat purchase (webhook)",                 "No",  "Called by RevenueCat, not app"),
    ("GET",    "/subscription/status",                     "Get current subscription status",                      "Yes", ""),
]

method_colors = {
    "GET":    ("E8F5E9", "1B5E20"),
    "POST":   ("E3F2FD", "0D47A1"),
    "PATCH":  ("FFF8E1", "E65100"),
    "DELETE": ("FFEBEE", "B71C1C"),
}

for i, row in enumerate(apis):
    r = i + 2
    data_row(ws7, r, row, alt=i % 2 == 1, height=20)
    mcell = ws7.cell(row=r, column=1)
    bg, fg = method_colors.get(row[0], ("FFFFFF", "000000"))
    mcell.fill = PatternFill("solid", fgColor=bg)
    mcell.font = Font(bold=True, color=fg)
    mcell.alignment = Alignment(horizontal="center", vertical="center")

set_col_widths(ws7, [10, 46, 42, 14, 36])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 8 — Data Models
# ═══════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Data Models")
ws8.sheet_view.showGridLines = False

header_style(ws8, 1, 5, ["Table", "Column", "Type", "Nullable", "Notes"])
ws8.row_dimensions[1].height = 28

schema = [
    # users
    ("users", "id",                   "UUID",        "No",  "Primary key, gen_random_uuid()"),
    ("users", "email",                "TEXT",        "No",  "Unique"),
    ("users", "password_hash",        "TEXT",        "Yes", "Null for OAuth-only users"),
    ("users", "display_name",         "TEXT",        "Yes", ""),
    ("users", "avatar_url",           "TEXT",        "Yes", ""),
    ("users", "tier",                 "TEXT",        "No",  "Default: 'free' | 'pro'"),
    ("users", "email_verified",       "BOOLEAN",     "No",  "Default: false"),
    ("users", "oauth_providers",      "JSONB",       "No",  "Array of {provider, provider_user_id}"),
    ("users", "settings",             "JSONB",       "No",  "theme, font_size, default_model, etc."),
    ("users", "custom_instructions",  "TEXT",        "Yes", "Max 2000 chars"),
    ("users", "history_saving",       "BOOLEAN",     "No",  "Default: true"),
    ("users", "created_at",           "TIMESTAMPTZ", "No",  "Default: now()"),
    ("users", "deleted_at",           "TIMESTAMPTZ", "Yes", "Soft delete for GDPR"),
    # conversations
    ("conversations", "id",           "UUID",        "No",  "Primary key"),
    ("conversations", "user_id",      "UUID",        "No",  "FK → users.id ON DELETE CASCADE"),
    ("conversations", "title",        "TEXT",        "No",  "Default: 'New Chat'"),
    ("conversations", "model_id",     "TEXT",        "No",  "Active model for this conversation"),
    ("conversations", "system_prompt","TEXT",        "Yes", "Per-conversation override"),
    ("conversations", "pinned",       "BOOLEAN",     "No",  "Default: false"),
    ("conversations", "archived",     "BOOLEAN",     "No",  "Default: false"),
    ("conversations", "created_at",   "TIMESTAMPTZ", "No",  "Default: now()"),
    ("conversations", "updated_at",   "TIMESTAMPTZ", "No",  "Updated on any message"),
    ("conversations", "last_message_at","TIMESTAMPTZ","Yes",""),
    # messages
    ("messages", "id",                "UUID",        "No",  "Primary key"),
    ("messages", "conversation_id",   "UUID",        "No",  "FK → conversations.id CASCADE"),
    ("messages", "role",              "TEXT",        "No",  "'user' | 'assistant' | 'system'"),
    ("messages", "content",           "TEXT",        "No",  ""),
    ("messages", "model_id",          "TEXT",        "Yes", "Set for assistant messages"),
    ("messages", "attachment_ids",    "UUID[]",      "No",  "Default: '{}'"),
    ("messages", "input_tokens",      "INTEGER",     "Yes", ""),
    ("messages", "output_tokens",     "INTEGER",     "Yes", ""),
    ("messages", "cancelled",         "BOOLEAN",     "No",  "Default: false"),
    ("messages", "created_at",        "TIMESTAMPTZ", "No",  "Default: now()"),
    # attachments
    ("attachments", "id",             "UUID",        "No",  "Primary key"),
    ("attachments", "user_id",        "UUID",        "No",  "FK → users.id CASCADE"),
    ("attachments", "type",           "TEXT",        "No",  "'image' | 'file'"),
    ("attachments", "original_name",  "TEXT",        "No",  ""),
    ("attachments", "mime_type",      "TEXT",        "No",  ""),
    ("attachments", "size_bytes",     "INTEGER",     "No",  ""),
    ("attachments", "storage_key",    "TEXT",        "No",  "S3/GCS object key"),
    ("attachments", "thumbnail_url",  "TEXT",        "Yes", ""),
    ("attachments", "created_at",     "TIMESTAMPTZ", "No",  "Default: now()"),
    # prompt_templates
    ("prompt_templates", "id",        "UUID",        "No",  "Primary key"),
    ("prompt_templates", "user_id",   "UUID",        "No",  "FK → users.id CASCADE"),
    ("prompt_templates", "name",      "TEXT",        "No",  "Max 20 templates per user"),
    ("prompt_templates", "content",   "TEXT",        "No",  ""),
    ("prompt_templates", "created_at","TIMESTAMPTZ", "No",  "Default: now()"),
    # refresh_tokens
    ("refresh_tokens", "id",          "UUID",        "No",  "Primary key"),
    ("refresh_tokens", "user_id",     "UUID",        "No",  "FK → users.id CASCADE"),
    ("refresh_tokens", "token_hash",  "TEXT",        "No",  "Unique; bcrypt hash"),
    ("refresh_tokens", "expires_at",  "TIMESTAMPTZ", "No",  "30 days from issue"),
    ("refresh_tokens", "revoked",     "BOOLEAN",     "No",  "Default: false"),
    ("refresh_tokens", "created_at",  "TIMESTAMPTZ", "No",  "Default: now()"),
]

table_colors = {
    "users":            "EDE7F6",
    "conversations":    "E3F2FD",
    "messages":         "E8F5E9",
    "attachments":      "FFF3E0",
    "prompt_templates": "FCE4EC",
    "refresh_tokens":   "F3E5F5",
}

prev_table = None
for i, row in enumerate(schema):
    r = i + 2
    data_row(ws8, r, row, height=18)
    tcell = ws8.cell(row=r, column=1)
    col = table_colors.get(row[0], "FFFFFF")
    tcell.fill = PatternFill("solid", fgColor=col)
    tcell.font = Font(bold=True)
    # Show table name only on first row of each table group
    if row[0] == prev_table:
        tcell.value = ""
    prev_table = row[0]

set_col_widths(ws8, [18, 22, 14, 10, 42])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 9 — Risks
# ═══════════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("Risks")
ws9.sheet_view.showGridLines = False

header_style(ws9, 1, 5, ["#", "Risk", "Likelihood", "Impact", "Mitigation"])
ws9.row_dimensions[1].height = 28

risks = [
    ("R-01","AI provider API pricing changes eat margin","Medium","High","Abstract provider layer; usage-based throttling; cost alerts"),
    ("R-02","App Store rejection (AI content policy)","Medium","High","Review Apple AI guidelines; add content filters; prepare appeal"),
    ("R-03","Streaming latency issues on slow networks","High","Medium","Graceful degradation; retry logic; offline message queuing"),
    ("R-04","RevenueCat webhook delays cause entitlement lag","Low","High","Optimistic entitlement grant with server confirmation"),
    ("R-05","GDPR/data breach notification requirement","Low","Very High","Minimal data retention; encryption at rest; incident response plan"),
    ("R-06","Provider outage (Anthropic/OpenAI/Google)","Medium","Medium","Per-provider error states; fallback model suggestion to user"),
    ("R-07","Flutter version fragmentation / breaking changes","Low","Medium","Pin Flutter version via FVM; staging env for upgrades"),
]

risk_colors = {"High": C_HIGH_RISK, "Very High": C_HIGH_RISK, "Medium": C_MED_RISK, "Low": C_LOW_RISK}

for i, row in enumerate(risks):
    r = i + 2
    data_row(ws9, r, row, alt=i % 2 == 1, height=28)
    # Color likelihood + impact
    for col_idx, val in [(3, row[2]), (4, row[3])]:
        cell = ws9.cell(row=r, column=col_idx)
        cell.fill = PatternFill("solid", fgColor=risk_colors.get(val, "FFFFFF"))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.font = Font(bold=True)

set_col_widths(ws9, [6, 44, 14, 14, 54])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 10 — Phased Delivery
# ═══════════════════════════════════════════════════════════════════════════
ws10 = wb.create_sheet("Phased Delivery")
ws10.sheet_view.showGridLines = False

header_style(ws10, 1, 3, ["Phase", "Feature / Activity", "Notes"])
ws10.row_dimensions[1].height = 28

phases = [
    # Phase 1
    ("Phase 1 — MVP\n(Months 1–3)", "Email + Google/Apple auth", "Core auth flow"),
    ("", "Text chat — Claude Haiku + Sonnet (streaming SSE)", ""),
    ("", "Text chat — GPT-4o mini + 4o (streaming SSE)", ""),
    ("", "Conversation CRUD + auto-title generation", ""),
    ("", "Markdown + syntax highlighting renderer", "Full flutter_highlight integration"),
    ("", "Dark/light mode + system default", ""),
    ("", "Free tier: 20 messages/day (Redis counter)", ""),
    ("", "Pro subscription via RevenueCat (monthly)", ""),
    ("", "Basic settings: theme, font size, default model", ""),
    ("Exit Criteria", "Internal beta on TestFlight + Play Console internal testing", ""),
    # Phase 2
    ("Phase 2 — Feature Complete\n(Months 4–5)", "Gemini Pro + Flash integration", "Provider gateway update"),
    ("", "Image uploads — camera + gallery picker", "flutter image_picker"),
    ("", "File uploads — PDF, DOCX, TXT", "Max 25 MB; server-side text extraction"),
    ("", "Voice input (STT)", "speech_to_text Flutter package"),
    ("", "Voice output (TTS)", "flutter_tts"),
    ("", "Custom instructions (global + per-conversation)", "Settings screen"),
    ("", "Data export + account deletion", "Async job → email link"),
    ("", "Conversation search (FTS + client-side)", "PostgreSQL tsvector"),
    ("", "Conversation rename, pin, archive", "Swipe + context menu"),
    ("", "Edit message & re-submit (with truncation confirm)", ""),
    ("Exit Criteria", "Feature freeze; QA begins; App Store/Play Store submission", ""),
    # Phase 3
    ("Phase 3 — Polish & Launch\n(Month 6)", "Performance optimization (cold start, render profiling)", ""),
    ("", "Accessibility audit (TalkBack + VoiceOver)", "WCAG AA target"),
    ("", "Localization — English + 2 additional languages", "i18n framework"),
    ("", "Analytics instrumentation", "Mixpanel or Amplitude"),
    ("", "Crash reporting setup", "Sentry or Firebase Crashlytics"),
    ("", "App Store submission (buffer: 1–3 days review)", ""),
    ("", "Play Store submission (buffer: 1–3 days review)", ""),
    ("", "Soft launch — limited geography first", "Validate metrics before global rollout"),
    ("Exit Criteria", "Live on both stores; D7 retention ≥ 30%", ""),
]

phase_fill = {
    "Phase 1 — MVP\n(Months 1–3)":               "E8EAF6",
    "Phase 2 — Feature Complete\n(Months 4–5)":   "E8F5E9",
    "Phase 3 — Polish & Launch\n(Month 6)":        "FFF3E0",
    "Exit Criteria": "FCE4EC",
}

current_phase_color = "E8EAF6"
for i, row in enumerate(phases):
    r = i + 2
    alt = i % 2 == 1
    data_row(ws10, r, row, alt=False, height=22)
    ph = row[0]
    if ph and ph in phase_fill:
        current_phase_color = phase_fill[ph]
    pcell = ws10.cell(row=r, column=1)
    pcell.fill = PatternFill("solid", fgColor=current_phase_color)
    pcell.font = Font(bold=True if ph else False, color="1A237E" if ph else "000000")
    pcell.alignment = Alignment(vertical="center", wrap_text=True)
    ws10.row_dimensions[r].height = 28 if ph else 22

set_col_widths(ws10, [32, 52, 40])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 11 — Open Decisions
# ═══════════════════════════════════════════════════════════════════════════
ws11 = wb.create_sheet("Open Decisions")
ws11.sheet_view.showGridLines = False

header_style(ws11, 1, 5, ["#", "Decision", "Owner", "Target Date", "Options / Notes"])
ws11.row_dimensions[1].height = 28

decisions = [
    ("OD-01", "App name and branding",      "Product",     "Month 1",      "TBD — needs trademark check"),
    ("OD-02", "Hosting provider",           "Engineering", "Month 1",      "Railway vs. Render vs. AWS ECS"),
    ("OD-03", "State management library",   "Engineering", "Month 1",      "Riverpod vs. BLoC"),
    ("OD-04", "Analytics platform",         "Product",     "Month 2",      "Mixpanel vs. Amplitude vs. PostHog"),
    ("OD-05", "Annual plan pricing",        "Product",     "Month 3",      "$79.99 (20% off) vs. $69.99 (30% off)"),
    ("OD-06", "Web search integration (v2)","Product",     "Post-launch",  "Tavily vs. Brave Search API"),
]

for i, row in enumerate(decisions):
    data_row(ws11, i + 2, row, alt=i % 2 == 1, height=22)

set_col_widths(ws11, [8, 34, 16, 16, 46])


# ═══════════════════════════════════════════════════════════════════════════
# Freeze top row on all sheets + set tab colors
# ═══════════════════════════════════════════════════════════════════════════
tab_colors = [
    ("Overview",               "3B4CB8"),
    ("Goals & Metrics",        "1565C0"),
    ("User Stories",           "6A1B9A"),
    ("Functional Requirements","00695C"),
    ("Non-Functional Reqs",    "558B2F"),
    ("Models & Tiers",         "E65100"),
    ("API Endpoints",          "0277BD"),
    ("Data Models",            "4527A0"),
    ("Risks",                  "B71C1C"),
    ("Phased Delivery",        "2E7D32"),
    ("Open Decisions",         "F57F17"),
]

for sheet, color in tab_colors:
    s = wb[sheet]
    s.freeze_panes = s["A2"]
    s.sheet_properties.tabColor = color

out_path = "/Users/bs1101/Work/1.Projects/claude-codex-chat/docs/PRD.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
