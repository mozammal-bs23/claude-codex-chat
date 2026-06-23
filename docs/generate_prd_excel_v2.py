import subprocess, sys

def install(pkg):
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

try:
    import openpyxl
except ImportError:
    install("openpyxl")
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colour palette (Puku brand — purple + orange accent) ────────────────────
C_HEADER_BG   = "4C2D8F"   # Puku purple
C_HEADER_FG   = "FFFFFF"
C_SUBHEAD_BG  = "EDE7F6"
C_SUBHEAD_FG  = "311B92"
C_ACCENT_BG   = "FF6B35"   # Puku orange accent
C_ACCENT_FG   = "FFFFFF"
C_ALT_ROW     = "F9F7FF"
C_MUST        = "B71C1C"
C_SHOULD      = "E65100"
C_COULD       = "2E7D32"
C_HIGH_RISK   = "FFCDD2"
C_MED_RISK    = "FFE0B2"
C_LOW_RISK    = "F1F8E9"
C_GREEN_FG    = "2E7D32"
C_RED_FG      = "C62828"
C_BORDER      = "D1C4E9"

# ── Shared helpers ───────────────────────────────────────────────────────────

def thin_border():
    t = Side(style="thin", color=C_BORDER)
    return Border(left=t, right=t, top=t, bottom=t)

def header_style(ws, row, cols, text_list, bg=C_HEADER_BG, fg=C_HEADER_FG, size=11):
    for i, txt in enumerate(text_list, 1):
        cell = ws.cell(row=row, column=i, value=txt)
        cell.font = Font(bold=True, color=fg, size=size)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()

def section_header(ws, row, ncols, text, bg=C_SUBHEAD_BG, fg=C_SUBHEAD_FG):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, color=fg, size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = thin_border()
    ws.row_dimensions[row].height = 22

def data_row(ws, row, values, alt=False, height=20, bold_col=None):
    fill = PatternFill("solid", fgColor=C_ALT_ROW) if alt else None
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border()
        if bold_col and i == bold_col:
            cell.font = Font(bold=True)
    ws.row_dimensions[row].height = height

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def color_cell(cell, hex_fg, hex_bg=None, bold=True, center=True):
    cell.font = Font(bold=bold, color=hex_fg)
    if hex_bg:
        cell.fill = PatternFill("solid", fgColor=hex_bg)
    cell.border = thin_border()
    if center:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def priority_color(p):
    if p in ("Must", "Must Have"):       return C_MUST
    if p in ("Should", "Should Have"):   return C_SHOULD
    if p in ("Could", "Could Have"):     return C_COULD
    return "444444"


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 — Overview
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Overview"
ws.sheet_view.showGridLines = False

header_style(ws, 1, 2, ["Puku Mobile App — PRD v2.0", ""], size=14)
ws.merge_cells("A1:B1")
ws.row_dimensions[1].height = 40

rows = [
    ("Document Version",      "2.0"),
    ("Date",                  "June 2026"),
    ("Status",                "Draft"),
    ("Source",                "Figma — https://scoop-white-12890271.figma.site/"),
    ("Product",               "Puku Mobile App (companion to Puku Editor)"),
    ("Tagline",               "The AI Code Editor That Understands Your Entire Codebase"),
    ("Platforms",             "iOS · Android"),
    ("AI Model",              "puku-ai-2.7 (Puku-hosted; not multi-provider)"),
    ("Account Type",          "Individual + Team / Organization"),
    ("Primary Navigation",    "Chats · Projects · Artifacts · Code"),
    ("Key Differentiators",   "Projects with Knowledge + Instructions · Code Sessions with Environments · GitHub integration · Incognito mode"),
    ("Tech Stack (Mobile)",   "Flutter (Dart) · BLoC/Cubit · GoRouter · FVM"),
    ("API Owner",             "Puku — provides and operates all backend APIs"),
    ("Mobile Team Scope",     "iOS + Android apps and Puku API integration only"),
    ("Figma Reference",       "https://scoop-white-12890271.figma.site/"),
]

for r, (k, v) in enumerate(rows, 2):
    c1 = ws.cell(row=r, column=1, value=k)
    c1.font = Font(bold=True, color=C_SUBHEAD_FG)
    c1.fill = PatternFill("solid", fgColor=C_SUBHEAD_BG)
    c1.alignment = Alignment(vertical="center")
    c1.border = thin_border()
    c2 = ws.cell(row=r, column=2, value=v)
    c2.alignment = Alignment(vertical="center", wrap_text=True)
    c2.border = thin_border()
    if r % 2 == 1:
        c2.fill = PatternFill("solid", fgColor=C_ALT_ROW)
    ws.row_dimensions[r].height = 22

set_col_widths(ws, [28, 70])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 — User Journeys
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("User Journeys")
ws2.sheet_view.showGridLines = False

header_style(ws2, 1, 4, ["ID", "Journey Name", "Narrative", "Success Outcome"])
ws2.row_dimensions[1].height = 28

journeys = [
    (
        "UJ-1",
        "Sign in and start a conversation",
        "A new user opens the app, sees the Puku Editor value proposition, signs in with Google or email, and within seconds starts typing in the chat input. Puku responds conversationally with awareness of the developer's coding context.",
        "User sends their first message within 90 seconds of first open.",
    ),
    (
        "UJ-2",
        "Organize work in a project",
        "A developer creates a project called 'Dev App', uploads a specification document to its Knowledge base, adds custom instructions telling Puku to behave as a senior iOS engineer, and starts a new chat scoped to that project.",
        "A project chat uses the uploaded knowledge and instructions in its responses.",
    ),
    (
        "UJ-3",
        "Run a live code session",
        "A developer opens the Code tab, taps 'New Session', selects puku-ai-2.7 and Node.js 20, connects GitHub, describes 'Build a REST API', and watches Puku execute. They toggle 'Accept edits automatically' to control change application.",
        "Code session completes and produces observable, reviewable artifacts.",
    ),
    (
        "UJ-4",
        "Continue a conversation privately",
        "A developer needs to explore a sensitive topic without saving to history. They open an Incognito chat from the lock icon, ask their question, then close — no history, no memory update.",
        "Incognito session leaves no trace in chat history or Puku memory.",
    ),
    (
        "UJ-5",
        "Review past work",
        "A developer opens the Chats screen, searches for an old conversation by keyword, finds it, and continues from where they left off.",
        "Full conversation history is searchable and continuable.",
    ),
]

for i, row in enumerate(journeys):
    data_row(ws2, i + 2, row, alt=i % 2 == 1, height=60)

set_col_widths(ws2, [8, 32, 70, 46])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3 — Screen Specifications
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Screen Specs")
ws3.sheet_view.showGridLines = False

header_style(ws3, 1, 5, ["#", "Screen", "Entry Point", "Key UI Elements", "Primary Actions"])
ws3.row_dimensions[1].height = 28

screens = [
    (
        "S-01", "Auth",
        "App cold open (unauthenticated)",
        "Dark theme · Purple gradient sphere · Puku 'P' logo · Headline with orange 'Codebase' · 'Continue with Google' white button · OR divider · Email input · Legal disclaimer",
        "Continue with Google · Enter email",
    ),
    (
        "S-02", "Chat (Home)",
        "Default after login · Nav drawer → New chat",
        "Light beige bg · Hamburger menu (left) · Lock icon for incognito (right) · Centered 'P' logo · Greeting 'Good afternoon, [username]' · Input bar: + | 'Chat with Puku...' | puku-ai-2.7 badge | 🎤 | Send",
        "Open nav drawer · Start incognito · Attach file · Select model · Voice input · Send message",
    ),
    (
        "S-03", "Nav Drawer",
        "Hamburger icon from any main screen",
        "Left overlay · 'Puku' title · Lock icons · Nav links: Chats / Projects / Artifacts / Code (with icons) · RECENTS section with recent chat titles · User avatar circle · New chat button (black) · Three-dot menu",
        "Navigate to Chats / Projects / Artifacts / Code · Open recent chat · Create new chat",
    ),
    (
        "S-04", "Incognito Chat",
        "Lock icon on Chat header",
        "Light bg · Header: Hamburger | 'Incognito chat' | X · Ghost icon (centered) · 'Incognito chats can't access memory…' info text · Admin visibility note with Learn more link · Same input bar as Chat",
        "Send incognito message · Close incognito (X)",
    ),
    (
        "S-05", "Chats (History)",
        "Nav drawer → Chats",
        "Header: Back | 'Chats' | Archive icon | Delete icon · Search bar 'Search Chats' · Chat list items: colored icon | title | date",
        "Search chats · Archive chat · Delete chat · Tap to open chat",
    ),
    (
        "S-06", "Projects",
        "Nav drawer → Projects",
        "Header: Hamburger | 'Projects' | Filter icon · Search bar 'Search projects' · Project list items: name | 'Edited [time]' · New project button (black, +)",
        "Search projects · Tap project to view details · New project",
    ),
    (
        "S-07", "Project Details",
        "Projects list → tap project",
        "Header: Back | project name | ⋮ · Subtitle · 'Created by Puku' icon · 'Private' lock badge · 'Project knowledge' card + 'Add knowledge' link (orange) · 'Custom instructions' card + 'Add instructions' link (orange) · Empty state: cloud + chat icon · New chat button (black, +)",
        "Add knowledge · Add instructions · Start new project chat · Edit/delete project (⋮)",
    ),
    (
        "S-08", "Project Knowledge",
        "Project Details → Project knowledge card",
        "Header: Back | 'Project Knowledge' | ⋮ · '0% of project capacity used' progress bar · Empty state: folder + plus icon · Description text · Add Content button (black)",
        "Add Content (opens bottom sheet) · View/delete uploaded files",
    ),
    (
        "S-09", "Add Content (Sheet)",
        "Project Knowledge → Add Content button",
        "Bottom sheet · X close · 'Add Content to Project' title · 'Upload files or create new text content' subtitle · 4 options with icons: Upload from device | Take picture | Choose image | Create new document",
        "Upload from device · Take picture · Choose image · Create new document",
    ),
    (
        "S-10", "Custom Instructions",
        "Project Details → Custom instructions card",
        "Header: Cancel | 'Custom Instructions' | Save (orange) · Description text · Large multiline textarea with placeholder example · Save Instructions button (disabled until filled)",
        "Enter instructions · Save · Cancel",
    ),
    (
        "S-11", "Code (Sessions List)",
        "Nav drawer → Code",
        "Header: Back | 'Code' · Empty state: terminal </> icon | 'No sessions found' | description text · New Session button (black, +)",
        "New Session · Tap existing session to resume",
    ),
    (
        "S-12", "New Session (Modal)",
        "Code screen → New Session button",
        "Modal · X close · MODEL dropdown: puku-ai-2.7 · ENVIRONMENT dropdown: Node.js 20 · SUGGESTED chips: Build REST API | Create React component | Write unit tests | Debug my code · Connect to GitHub button · 'Accept edits automatically' toggle (OFF) · Input bar: + | 'Describe what you want to build...' | model badge | 🎤 | Send (gray/disabled)",
        "Select model · Select environment · Tap suggested chip · Connect GitHub · Toggle accept-edits · Describe task → Send",
    ),
    (
        "S-13", "Create Project (Modal)",
        "Projects screen → New project button",
        "Modal · X close · 'Create a project' title · 'What are you working on?' → Name input · 'What are you trying to achieve?' → Description textarea · Visibility: Private (lock, orange border) / Organization (building icon) · Create project button (disabled until name entered)",
        "Enter name · Enter description · Select visibility · Create project",
    ),
    (
        "S-14", "Settings",
        "Nav drawer → user avatar / settings area",
        "Header: Hamburger | 'Settings' | ℹ️ · User: email | @username | Team badge (black) · List items with icon + label + optional subtitle + chevron: Profile | Billing | Usage | Capabilities (2 enabled) | Permissions | Font style (Default) | Voice | Haptic feedback (toggle) | Notifications | Shared links · Log out (red)",
        "Tap any row to navigate to sub-screen · Toggle haptic feedback · Log out",
    ),
]

for i, row in enumerate(screens):
    data_row(ws3, i + 2, row, alt=i % 2 == 1, height=50)

set_col_widths(ws3, [6, 22, 30, 70, 48])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4 — Functional Requirements
# ═══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Functional Requirements")
ws4.sheet_view.showGridLines = False

header_style(ws4, 1, 5, ["FR-ID", "Screen", "Requirement", "Priority", "Acceptance Criteria (summary)"])
ws4.row_dimensions[1].height = 28

func_reqs = [
    # Auth
    ("FR-01", "S-01 Auth",        "Google OAuth sign-in — tap 'Continue with Google', complete platform OAuth flow, receive Puku session token",                                               "Must Have",   "Auth token stored in secure storage; navigates to Chat on success"),
    ("FR-02", "S-01 Auth",        "Email sign-in/sign-up — enter email, proceed through Puku's auth flow (OTP or password, per Puku API)",                                                   "Must Have",   "Valid email proceeds; invalid email shows error"),
    ("FR-03", "S-01 Auth",        "Session persistence — auth token survives app restart; silent token refresh before expiry",                                                                "Must Have",   "User not prompted to re-authenticate on normal restart"),
    ("FR-04", "S-01 Auth",        "Sign-out — clear local tokens and cached data; return to Auth screen",                                                                                     "Must Have",   "Settings → Log out clears all local state"),
    # Chat
    ("FR-05", "S-02 Chat",        "Streaming chat — user sends message; Puku responds via SSE stream; tokens appear in real time",                                                            "Must Have",   "First token visible within 1.5 s p95; no frame drops during stream"),
    ("FR-06", "S-02 Chat",        "Model selector badge — tap 'puku-ai-2.7' badge to open model picker; selected model is used for next message",                                             "Must Have",   "Model list sourced from Puku API; badge updates after selection"),
    ("FR-07", "S-02 Chat",        "Attachment picker (+) — open sheet with image and file options; selected files appear as preview strip above input bar",                                   "Must Have",   "Selected attachment visible; removed via long-press/tap X"),
    ("FR-08", "S-02 Chat",        "Voice input — tap microphone; record speech; transcription populates input field; user edits before sending",                                              "Should Have", "Transcription appears; user can edit before send"),
    ("FR-09", "S-02 Chat",        "Stop streaming — cancel button appears during stream; tap stops response mid-stream; partial text preserved",                                               "Must Have",   "Cancel button visible while streaming; partial response saved"),
    ("FR-10", "S-02 Chat",        "Message rendering — markdown, code blocks with syntax highlighting and copy button, bold, italic, lists, tables",                                          "Must Have",   "All markdown elements render correctly; code copy shows 'Copied!' toast"),
    # Incognito
    ("FR-11", "S-04 Incognito",   "Incognito chat — open from lock icon; messages sent without server persistence; session not added to history",                                             "Must Have",   "Incognito chat absent from Chats list after closing"),
    ("FR-12", "S-04 Incognito",   "No memory update — incognito messages do not update Puku memory or affect future non-incognito responses",                                                 "Must Have",   "Confirmed via Puku API no-save mode"),
    ("FR-13", "S-04 Incognito",   "Admin disclaimer — 'Chat history is still visible to your admin' note always visible inside incognito",                                                   "Must Have",   "Disclaimer present at all times in incognito"),
    # Chat History
    ("FR-14", "S-05 Chats",       "Chat list — all saved chats sorted newest first; show title, icon color, and date",                                                                        "Must Have",   "List loads within 1 s; correct order"),
    ("FR-15", "S-05 Chats",       "Search chats — filter list by title and message content; real-time as user types",                                                                         "Should Have", "Results update within 300 ms of keystroke"),
    ("FR-16", "S-05 Chats",       "Archive chat — header archive icon moves selected chat(s) to archived state; not shown in main list",                                                     "Should Have", "Archived chats accessible from archive view"),
    ("FR-17", "S-05 Chats",       "Delete chat — header delete icon removes selected chat(s) permanently after confirmation",                                                                 "Must Have",   "Confirmation dialog shown before delete"),
    # Projects
    ("FR-18", "S-06 Projects",    "Projects list — show all user projects with name and last-edited timestamp; newest first",                                                                  "Must Have",   "List loads within 1 s"),
    ("FR-19", "S-06 Projects",    "Search projects — filter list by project name in real time",                                                                                               "Should Have", "Results update within 300 ms"),
    ("FR-20", "S-06 Projects",    "Create project — New project button opens Create Project modal; on submit, project appears in list",                                                       "Must Have",   "Project visible immediately after creation"),
    # Project Details
    ("FR-21", "S-07 Proj Detail", "Project overview — display name, subtitle, created-by, and privacy badge (Private / Organization)",                                                       "Must Have",   "Data fetched from Puku API"),
    ("FR-22", "S-07 Proj Detail", "Project knowledge entry — 'Add knowledge' link navigates to Project Knowledge screen",                                                                    "Must Have",   "Navigates to S-08"),
    ("FR-23", "S-07 Proj Detail", "Custom instructions entry — 'Add instructions' link navigates to Custom Instructions screen",                                                              "Must Have",   "Navigates to S-10"),
    ("FR-24", "S-07 Proj Detail", "Project chats — list of chats associated with this project; empty state shown until first chat",                                                           "Must Have",   "Project-scoped chats appear here"),
    ("FR-25", "S-07 Proj Detail", "New project chat — 'New chat' button starts a chat with this project's knowledge + instructions applied",                                                  "Must Have",   "New chat opens with project context active"),
    # Knowledge
    ("FR-26", "S-08 Knowledge",   "Capacity indicator — show '0% of project capacity used' progress bar; update after uploads",                                                               "Must Have",   "Reflects Puku API response"),
    ("FR-27", "S-08 Knowledge",   "File list — show uploaded files with name, type, size after upload; replace empty state",                                                                  "Must Have",   "Files appear immediately after upload"),
    ("FR-28", "S-08 Knowledge",   "Delete file — allow removing a specific knowledge file via swipe or long-press",                                                                           "Should Have", "File removed and capacity updated"),
    # Add Content
    ("FR-29", "S-09 Add Content", "Upload from device — open file picker; support PDF, DOCX, TXT, CSV, XLSX, and common code files",                                                         "Must Have",   "File picker opens; unsupported types show error"),
    ("FR-30", "S-09 Add Content", "Take picture — open device camera; captured image uploaded to project knowledge",                                                                          "Should Have", "Camera opens; image added to knowledge"),
    ("FR-31", "S-09 Add Content", "Choose image — open photo library; selected image uploaded to project knowledge",                                                                          "Should Have", "Library opens; image added to knowledge"),
    ("FR-32", "S-09 Add Content", "Create new document — in-app text editor for pasting/typing content; save as knowledge file",                                                              "Should Have", "Text saved as knowledge file"),
    # Custom Instructions
    ("FR-33", "S-10 Instructions","Save custom instructions — enter free-form text; tap Save; instructions stored against project and applied to all project chats",                         "Must Have",   "Instructions present in system prompt of next project chat"),
    ("FR-34", "S-10 Instructions","Cancel without saving — tap Cancel; unsaved changes discarded; confirmation if text was entered",                                                          "Should Have", "Unsaved changes discarded; prompt shown if dirty"),
    # Code Sessions
    ("FR-35", "S-11 Code",        "Sessions list — list active and historical code sessions with name, environment, last-active time; empty state when none",                                "Must Have",   "List from Puku API; empty state shown correctly"),
    ("FR-36", "S-11 Code",        "Resume session — tap existing session to open it",                                                                                                         "Must Have",   "Session state loaded from Puku API"),
    # New Session
    ("FR-37", "S-12 New Session", "Model selection — dropdown shows Puku-API-sourced model list; puku-ai-2.7 default",                                                                        "Must Have",   "Model list from config API; selected model used for session"),
    ("FR-38", "S-12 New Session", "Environment selection — dropdown shows Puku-API-sourced runtimes (Node.js 20, etc.)",                                                                     "Must Have",   "Environment list from config API"),
    ("FR-39", "S-12 New Session", "Suggested chips — tapping a chip populates the description input field",                                                                                    "Should Have", "Input field filled with chip text"),
    ("FR-40", "S-12 New Session", "GitHub connection — OAuth flow to connect GitHub; persists in Puku account once authorized",                                                               "Must Have",   "GitHub connected state shown after auth; persists across sessions"),
    ("FR-41", "S-12 New Session", "'Accept edits automatically' toggle — OFF by default; when ON, Puku applies code changes without confirmation; persisted per session",                    "Must Have",   "Toggle state saved; affects Puku session permission mode"),
    ("FR-42", "S-12 New Session", "Start session — enter description; tap Send (enabled only with text); navigate into active session view",                                                  "Must Have",   "Session created via Puku API; session view opens"),
    # Create Project
    ("FR-43", "S-13 Create Proj", "Project name required — 'Create project' button disabled until name field has content",                                                                    "Must Have",   "Button state matches input state"),
    ("FR-44", "S-13 Create Proj", "Visibility selection — Private (default) or Organization; options reflect user's org membership from Puku API",                                           "Must Have",   "Org name shown if user has an org"),
    ("FR-45", "S-13 Create Proj", "Create project — submit creates project via Puku API; navigate to Project Details; project appears in Projects list",                                      "Must Have",   "Project visible in list immediately"),
    # Settings
    ("FR-46", "S-14 Settings",    "Profile display — show email, @username, and Team badge from Puku API",                                                                                   "Must Have",   "Data matches Puku account"),
    ("FR-47", "S-14 Settings",    "Settings navigation — tapping any row navigates to its detail screen (Profile, Billing, Usage, Capabilities, Permissions, Font style, Voice, Notifications, Shared links)", "Must Have", "All rows navigable"),
    ("FR-48", "S-14 Settings",    "Haptic feedback toggle — takes effect immediately on toggle; persisted in user preferences",                                                               "Should Have", "Toggle state persists across app restarts"),
    ("FR-49", "S-14 Settings",    "Log out — clears all local tokens and cached data; returns to Auth screen",                                                                                "Must Have",   "Auth screen shown; no user data accessible after logout"),
]

for i, row in enumerate(func_reqs):
    r = i + 2
    alt = i % 2 == 1
    data_row(ws4, r, row, alt=alt, height=28)
    pcell = ws4.cell(row=r, column=4)
    color_cell(pcell, priority_color(row[3]), center=True, bold=True)

set_col_widths(ws4, [8, 18, 62, 14, 46])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 5 — API Contracts
# ═══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("API Contracts")
ws5.sheet_view.showGridLines = False

header_style(ws5, 1, 5, ["Feature Area", "Required Puku API Capability", "Protocol", "Mobile Feature(s)", "Status"])
ws5.row_dimensions[1].height = 28

apis = [
    ("Authentication",    "OAuth / OIDC or equivalent mobile-safe auth flow",                                      "HTTPS",        "S-01 Auth",                         "Needs confirmation"),
    ("Authentication",    "Short-lived access token + refresh token; logout invalidation",                          "HTTPS",        "S-01 Auth, S-14 Log out",           "Needs confirmation"),
    ("User Profile",      "GET /users/me — email, @username, team/org membership, plan",                            "HTTPS / REST", "S-14 Settings",                     "Needs confirmation"),
    ("Chat",              "POST message with SSE streaming response",                                                "SSE",          "S-02 Chat, S-04 Incognito",         "Needs confirmation"),
    ("Chat",              "GET conversation list — paginated, with title and last-active",                           "HTTPS / REST", "S-05 Chats",                        "Needs confirmation"),
    ("Chat",              "GET/PATCH/DELETE single conversation",                                                    "HTTPS / REST", "S-05 Chats",                        "Needs confirmation"),
    ("Chat",              "Full-text search conversations",                                                          "HTTPS / REST", "S-05 Chats search",                 "Needs confirmation"),
    ("Chat — Incognito",  "Send message without persistence (server-side no-save mode)",                             "SSE",          "S-04 Incognito",                    "Needs confirmation"),
    ("Models",            "Config endpoint returning available model list with capabilities",                        "HTTPS / REST", "S-02 model badge, S-12 New Session","Needs confirmation"),
    ("Projects",          "CRUD endpoints for project entities (name, description, visibility)",                     "HTTPS / REST", "S-06 Projects, S-07 Proj Details",  "Needs confirmation"),
    ("Project Knowledge", "Upload document / image to project knowledge base",                                       "HTTPS / REST", "S-08 Knowledge, S-09 Add Content",  "Needs confirmation"),
    ("Project Knowledge", "Capacity and usage metadata per project",                                                 "HTTPS / REST", "S-08 Knowledge",                    "Needs confirmation"),
    ("Project Knowledge", "List and delete knowledge files per project",                                             "HTTPS / REST", "S-08 Knowledge",                    "Needs confirmation"),
    ("Instructions",      "Store and retrieve per-project custom instruction text",                                  "HTTPS / REST", "S-10 Custom Instructions",          "Needs confirmation"),
    ("Code Sessions",     "List active and historical sessions with model, environment, state",                      "HTTPS / REST", "S-11 Code",                         "Needs confirmation"),
    ("Code Sessions",     "Create session with model, environment, description, GitHub repo, accept-edits flag",    "HTTPS / REST", "S-12 New Session",                  "Needs confirmation"),
    ("Code Sessions",     "Real-time event stream — agent progress, file edits, command output, state changes",     "SSE / WS",     "S-12 Active Session",               "Needs confirmation"),
    ("Environments",      "Config endpoint returning available runtime environments (Node.js 20, Python, etc.)",    "HTTPS / REST", "S-12 New Session environment picker","Needs confirmation"),
    ("GitHub",            "OAuth flow to connect GitHub account; persist in Puku account",                          "OAuth",        "S-12 New Session",                  "Needs confirmation"),
    ("Usage & Quota",     "Usage metrics, quota remaining, credit consumption",                                      "HTTPS / REST", "S-14 Settings → Usage",             "Needs confirmation"),
    ("Capabilities",      "Feature flags / capabilities enabled for this account",                                   "HTTPS / REST", "S-14 Settings → Capabilities",      "Needs confirmation"),
    ("Notifications",     "Push token registration; notification preference management",                             "HTTPS / REST", "Push, S-14 Notifications",          "Needs confirmation"),
    ("Shared Links",      "CRUD for publicly shared chat or code session links",                                     "HTTPS / REST", "S-14 Settings → Shared links",      "Needs confirmation"),
    ("Org / Team",        "Organization membership and org-visible projects",                                        "HTTPS / REST", "S-13 visibility, S-14 Team badge",  "Needs confirmation"),
]

status_colors = {
    "Needs confirmation": ("FFF3E0", "E65100"),
    "Confirmed":          ("E8F5E9", "2E7D32"),
    "Blocked":            ("FFEBEE", "B71C1C"),
}

for i, row in enumerate(apis):
    r = i + 2
    data_row(ws5, r, row, alt=i % 2 == 1, height=24)
    scell = ws5.cell(row=r, column=5)
    bg, fg = status_colors.get(row[4], ("FFFFFF", "000000"))
    scell.fill = PatternFill("solid", fgColor=bg)
    scell.font = Font(bold=True, color=fg)
    scell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

set_col_widths(ws5, [22, 56, 14, 36, 20])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 6 — Non-Functional Requirements
# ═══════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Non-Functional Reqs")
ws6.sheet_view.showGridLines = False

header_style(ws6, 1, 4, ["ID", "Category", "Requirement", "Target / Detail"])
ws6.row_dimensions[1].height = 28

nfr = [
    # Performance
    ("NFR-01", "Performance",    "App cold start → first screen visible",          "< 3 seconds on mid-range Android"),
    ("NFR-02", "Performance",    "Chat / Projects screen loads (cached state)",     "< 2 seconds"),
    ("NFR-03", "Performance",    "First streaming token after message send",        "< 1.5 seconds (p95)"),
    ("NFR-04", "Performance",    "Event render latency after client receipt",       "< 250 ms"),
    ("NFR-05", "Performance",    "Offline mode — cached content readable",          "Recent chats and project metadata readable; marked stale"),
    # Security
    ("NFR-06", "Security",       "All network traffic",                             "TLS 1.3 minimum"),
    ("NFR-07", "Security",       "Auth token storage",                              "flutter_secure_storage — Keychain (iOS) / EncryptedSharedPreferences (Android)"),
    ("NFR-08", "Security",       "Puku credentials",                               "Short-lived, scoped, revocable; never hardcoded in app bundle"),
    ("NFR-09", "Security",       "Code session edits — accept-edits OFF by default","User must explicitly enable; OFF is the secure default"),
    ("NFR-10", "Security",       "Telemetry and support bundles",                   "Source code, prompts, secrets excluded by default"),
    # Reliability
    ("NFR-11", "Reliability",    "Crash-free sessions",                             "≥ 99.5%"),
    ("NFR-12", "Reliability",    "Chat and session state after reconnect",           "Reconciled from Puku API; no duplicate message sends"),
    ("NFR-13", "Reliability",    "Push notifications for code session events",       "Delivered within 10 seconds for ≥ 99% of accepted deliveries"),
    # Accessibility
    ("NFR-14", "Accessibility",  "Colour contrast",                                 "WCAG 2.2 AA on all user-facing screens"),
    ("NFR-15", "Accessibility",  "Screen reader support",                           "VoiceOver (iOS) + TalkBack (Android)"),
    ("NFR-16", "Accessibility",  "Touch targets",                                   "Minimum 44 × 44 pt"),
    ("NFR-17", "Accessibility",  "Dynamic text",                                    "Respect system font size setting"),
    ("NFR-18", "Accessibility",  "Reduced motion",                                  "Respect system reduced-motion preference"),
    # Compatibility
    ("NFR-19", "Compatibility",  "iOS minimum version",                             "TBD — see Open Questions OQ-1"),
    ("NFR-20", "Compatibility",  "Android minimum version",                         "TBD — see Open Questions OQ-1"),
    ("NFR-21", "Compatibility",  "Flutter toolchain",                               "Pinned via FVM; all commands use fvm flutter / fvm dart"),
    # Code quality (Flutter project rules)
    ("NFR-22", "Code Quality",   "State management",                                "BLoC / Cubit exclusively; no ChangeNotifier or setState for business state"),
    ("NFR-23", "Code Quality",   "Navigation",                                      "GoRouter exclusively; no Navigator.push/pop"),
    ("NFR-24", "Code Quality",   "UI class size",                                   "Max 100 lines per widget class; extract to dedicated widget classes"),
    ("NFR-25", "Code Quality",   "Sizes",                                           "AppSize constants only — no hardcoded numbers in UI"),
    ("NFR-26", "Code Quality",   "Colors and text styles",                          "Theme only — context.primary, context.customTextStyles.* etc."),
]

for i, row in enumerate(nfr):
    data_row(ws6, i + 2, row, alt=i % 2 == 1, height=22)

set_col_widths(ws6, [10, 18, 46, 52])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 7 — MVP Scope
# ═══════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("MVP Scope")
ws7.sheet_view.showGridLines = False

# In Scope
section_header(ws7, 1, 3, "✅  IN SCOPE — MVP v1", bg="E8F5E9", fg="1B5E20")
header_style(ws7, 2, 3, ["#", "Feature", "Notes"], bg="2E7D32")
ws7.row_dimensions[2].height = 24

in_scope = [
    ("1",  "iOS and Android apps",                                          "Flutter"),
    ("2",  "Authentication — Google OAuth + email sign-in",                  "Via Puku auth flow"),
    ("3",  "Chat — streaming AI conversation",                               "SSE; puku-ai model"),
    ("4",  "Chat — model selection badge",                                   "From Puku config API"),
    ("5",  "Chat — attachments (image + file upload)",                       "+  button in input bar"),
    ("6",  "Chat — voice input",                                             "STT → text field"),
    ("7",  "Incognito chat",                                                  "No history; no memory update"),
    ("8",  "Navigation drawer (Chats / Projects / Artifacts / Code)",        "Artifacts placeholder OK"),
    ("9",  "Chat history — list, search, archive, delete",                   ""),
    ("10", "Projects — create, list, view details",                          ""),
    ("11", "Project knowledge — upload files / images / text",               "Capacity indicator"),
    ("12", "Custom instructions — per-project system prompt",                ""),
    ("13", "Code sessions — create with model + environment",                "Node.js 20 + others per API"),
    ("14", "Code sessions — GitHub OAuth connection",                        ""),
    ("15", "Code sessions — accept-edits toggle",                            "OFF by default"),
    ("16", "Settings — profile, billing, usage, capabilities, permissions",  ""),
    ("17", "Settings — font style, voice, haptic, notifications, shared links",""),
    ("18", "Settings — log out",                                             ""),
    ("19", "Push notifications for code session events",                     "Via Puku backend"),
]

for i, row in enumerate(in_scope):
    data_row(ws7, i + 3, row, alt=i % 2 == 1, height=22)

# Spacer
ws7.row_dimensions[len(in_scope) + 3].height = 12

# Out of Scope
out_row = len(in_scope) + 4
section_header(ws7, out_row, 3, "❌  OUT OF SCOPE — v1", bg="FFEBEE", fg="B71C1C")
header_style(ws7, out_row + 1, 3, ["#", "Feature", "Notes"], bg="C62828")
ws7.row_dimensions[out_row + 1].height = 24

out_scope = [
    ("1",  "Artifacts section content",               "Nav slot present; content deferred to v2"),
    ("2",  "Full code diff review UI",                "Deferred to v2"),
    ("3",  "Multi-agent / team orchestration",        "Deferred"),
    ("4",  "On-device code execution",                "Puku cloud environments only"),
    ("5",  "Web app",                                  "Mobile-only for v1"),
    ("6",  "iPad / tablet optimized layout",          "Responsive but not optimized"),
    ("7",  "RTL language support",                     "Deferred post-launch"),
    ("8",  "Admin / SSO / enterprise policy console", "Not in MVP"),
    ("9",  "Voice output (TTS)",                       "Could Have — defer if timeline tight"),
    ("10", "Annual billing plan",                      "Monthly only in v1"),
]

for i, row in enumerate(out_scope):
    data_row(ws7, out_row + 2 + i, row, alt=i % 2 == 1, height=22)

set_col_widths(ws7, [6, 50, 42])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 8 — Open Questions
# ═══════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Open Questions")
ws8.sheet_view.showGridLines = False

header_style(ws8, 1, 5, ["#", "Question", "Owner", "Required By", "Status"])
ws8.row_dimensions[1].height = 28

questions = [
    ("OQ-1",  "Minimum iOS and Android versions",                                                                               "Engineering + Puku", "Sprint 1",  "Open"),
    ("OQ-2",  "Puku authentication flow — OAuth/OIDC spec, token TTLs, refresh, MFA, logout",                                  "Puku",               "Sprint 1",  "Open"),
    ("OQ-3",  "Real-time protocol — SSE vs. WebSocket for chat and code session events; event replay / cursor contract",         "Puku",               "Sprint 1",  "Open"),
    ("OQ-4",  "Available model list — is puku-ai-2.7 the only model at launch? Format of config endpoint?",                     "Puku",               "Sprint 1",  "Open"),
    ("OQ-5",  "Organization / team API — how is the Team badge populated? Which org entities are exposed?",                      "Puku",               "Sprint 1",  "Open"),
    ("OQ-6",  "Analytics platform — does Puku require a specific SDK (Mixpanel, Amplitude, PostHog)?",                          "Puku / Product",     "Sprint 1",  "Open"),
    ("OQ-7",  "Available environments for Code sessions beyond Node.js 20 (Python, Go, Rust, etc.)",                            "Puku",               "Sprint 2",  "Open"),
    ("OQ-8",  "Knowledge base — capacity limits, supported file types, max file size",                                          "Puku",               "Sprint 2",  "Open"),
    ("OQ-9",  "GitHub OAuth — Puku-owned flow or mobile-owned? Does connection persist in Puku account?",                       "Puku",               "Sprint 2",  "Open"),
    ("OQ-10", "Notification types — which session events trigger push notifications from Puku backend?",                         "Puku",               "Sprint 2",  "Open"),
    ("OQ-11", "Artifacts — what is an artifact in Puku's data model? What should the Artifacts screen show?",                   "Puku + Product",     "Sprint 2",  "Open"),
    ("OQ-12", "Billing and Usage screens — does Puku provide a web URL to embed or do we build native UI?",                     "Puku",               "Sprint 3",  "Open"),
    ("OQ-13", "Shared links — what content can be shared and how are public links managed?",                                     "Puku",               "Sprint 3",  "Open"),
    ("OQ-14", "App name in stores — 'Puku' or 'Puku Editor'? Trademark status?",                                               "Puku / Legal",       "Pre-submission", "Open"),
]

status_colors_q = {
    "Open":     ("FFF8E1", "E65100"),
    "Resolved": ("E8F5E9", "2E7D32"),
    "Blocked":  ("FFEBEE", "B71C1C"),
}

for i, row in enumerate(questions):
    r = i + 2
    data_row(ws8, r, row, alt=i % 2 == 1, height=28)
    scell = ws8.cell(row=r, column=5)
    bg, fg = status_colors_q.get(row[4], ("FFFFFF", "000000"))
    scell.fill = PatternFill("solid", fgColor=bg)
    scell.font = Font(bold=True, color=fg)
    scell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

set_col_widths(ws8, [8, 68, 22, 18, 14])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 9 — Success Metrics
# ═══════════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("Success Metrics")
ws9.sheet_view.showGridLines = False

header_style(ws9, 1, 4, ["ID", "Metric", "Target", "Type"])
ws9.row_dimensions[1].height = 28

metrics = [
    # Onboarding
    ("SM-1",  "Time from first open → first message sent",                  "≤ 90 seconds",              "Onboarding"),
    ("SM-2",  "New-user time from open → first code session submitted",     "≤ 10 minutes",              "Onboarding"),
    # Engagement
    ("SM-3",  "Returning-user time from app open → submitted task",         "≤ 30 seconds",              "Engagement"),
    ("SM-4",  "Weekly active users completing ≥ 1 code session",            "≥ 35% of WAU",              "Engagement"),
    ("SM-5",  "Projects with ≥ 1 knowledge file added",                     "≥ 40% of created projects", "Engagement"),
    # Technical
    ("SM-6",  "First streaming token latency (p95)",                        "< 1.5 seconds",             "Technical"),
    ("SM-7",  "Puku event receipt to mobile render",                        "p95 ≤ 1 second",            "Technical"),
    ("SM-8",  "Crash-free mobile sessions",                                  "≥ 99.5%",                   "Technical"),
    ("SM-9",  "Approval notifications delivered within 10 seconds",         "≥ 99% of accepted deliveries","Technical"),
    # Quality
    ("SM-10", "App Store / Play Store rating",                               "≥ 4.4 stars",               "Quality"),
    ("SM-11", "Workspace load success rate",                                 "≥ 95%",                     "Quality"),
]

section_header(ws9, len(metrics) + 3, 4, "Counter-metrics (watch for anti-patterns)", bg="FFF3E0", fg="E65100")

counter_metrics = [
    ("SM-C1", "Unintended or incorrectly scoped destructive code actions",  "0 known incidents",         "Safety"),
    ("SM-C2", "Code sessions shown as successful when required checks failed","< 1%",                    "Safety"),
    ("SM-C3", "Users granting broad persistent GitHub rules in first week",  "Growth requires security review", "Safety"),
]

type_colors = {
    "Onboarding": ("E3F2FD", "0D47A1"),
    "Engagement": ("E8F5E9", "1B5E20"),
    "Technical":  ("F3E5F5", "4A148C"),
    "Quality":    ("FFF8E1", "F57F17"),
    "Safety":     ("FFEBEE", "B71C1C"),
}

for i, row in enumerate(metrics):
    r = i + 2
    data_row(ws9, r, row, alt=i % 2 == 1, height=22)
    tcell = ws9.cell(row=r, column=4)
    bg, fg = type_colors.get(row[3], ("FFFFFF", "000000"))
    tcell.fill = PatternFill("solid", fgColor=bg)
    tcell.font = Font(bold=True, color=fg)
    tcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for i, row in enumerate(counter_metrics):
    r = len(metrics) + 4 + i
    data_row(ws9, r, row, alt=i % 2 == 1, height=22)
    tcell = ws9.cell(row=r, column=4)
    bg, fg = type_colors.get(row[3], ("FFFFFF", "000000"))
    tcell.fill = PatternFill("solid", fgColor=bg)
    tcell.font = Font(bold=True, color=fg)
    tcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

set_col_widths(ws9, [8, 62, 32, 16])


# ═══════════════════════════════════════════════════════════════════════════
# Freeze panes + tab colors + hide gridlines
# ═══════════════════════════════════════════════════════════════════════════
tab_colors = [
    ("Overview",               "4C2D8F"),  # Puku purple
    ("User Journeys",          "6A1B9A"),
    ("Screen Specs",           "0277BD"),
    ("Functional Requirements","00695C"),
    ("API Contracts",          "1565C0"),
    ("Non-Functional Reqs",    "558B2F"),
    ("MVP Scope",              "2E7D32"),
    ("Open Questions",         "E65100"),  # Puku orange
    ("Success Metrics",        "880E4F"),
]

for sheet_name, color in tab_colors:
    s = wb[sheet_name]
    s.freeze_panes = s["A2"]
    s.sheet_properties.tabColor = color
    s.sheet_view.showGridLines = False

out_path = "/Users/bs1101/Work/1.Projects/claude-codex-chat/docs/PRD_v2.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
